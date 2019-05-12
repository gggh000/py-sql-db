#!/usr/bin/python3

'''
PASSWORD AND PHONE MANAGEMENT SYSTEM.
GUYEN GN.
THIS UTILITY REQUIRES MYSQL SERVER RUNNING ON CONFIG_MYSQL_IP
If repeatedly run this will re-import the xlsx without checking over and over again.
If you need to reimport it, then it is advisabae to remove the table first by:

mysql -i root -p
show databases;
use passwds
show tables;
drop table <MY_SQL_DB_PWS_TABLE>;
exit

'''

#   Helper functions.

def printBarDouble():
    print("=================================================")

def printBarSingle():
    print("-------------------------------------------------")

#   Import statements. 

import os
import sys
import re
import mysql.connector
import time
from pySqlDbApi import *
from openpyxl import *
from cmnLib3 import *

presenceDb = None
presenceTable = None
insertErrors = {}
debug = 0

os.system("clear")

#    Initialize variables.

dbNamePws=os.environ["MYSQL_DB_PWS"]
dbNamePwsTbl=os.environ["MYSQL_DB_PWS_TABLE"]
dbNameUser=os.environ["MYSQL_DB_USER"]
mySqlServerIp=os.environ["MYSQL_SERVER_IP"]
print("database name: ", dbNamePws)
fileNameExcel='NUMBERS-EXPORT.xlsx'

#    Connect to mysql server, root without password.

pwManager = mysqlManager()

pwManager.mainMenuSelectTbl()

#   declare excel output to export to.

wb = Workbook()
ws = wb.active
ws.title = "NUMBERS-EXPORT"

counter = 0;

pwManager.mycursor.execute("select * from " + str(pwManager.tableToUse))

counter = 0
col_ranges=['A','B','C','D','E','F','G','H']
ws1 = wb.create_sheet("passwords")

for x in pwManager.mycursor:
    counter += 1
    print(x)

    if debug:
        print("len(x): ", len(x))
        print("len(col_ranges): ", len(col_ranges))
    
    for i in range(len(x)):
        coord = col_ranges[i] + str(counter)

        if debug:
            print("coord: ", coord)

        ws[coord] = x[i]

wb.save('pw.xlsx')

